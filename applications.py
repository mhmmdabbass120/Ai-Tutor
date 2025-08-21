"""
Advanced Applications Module for AI Tutor
Handles file processing, PDF generation, image analysis, and advanced features
"""

import streamlit as st
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import plotly.express as px
import plotly.graph_objects as go
from PIL import Image
import io
import base64
from datetime import datetime
import json
import os
from typing import Dict, List, Tuple, Any, Optional
import re

# PDF and Document Processing
try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.platypus.flowables import Image as RLImage
    from fpdf import FPDF
    import PyPDF2
    import pdfplumber
except ImportError:
    st.warning("Some PDF libraries not available. Install: pip install reportlab PyPDF2 pdfplumber fpdf2")

# OCR and Image Processing
try:
    import pytesseract
    import cv2
except ImportError:
    st.warning("OCR libraries not available. Install: pip install pytesseract opencv-python")

# Document Processing
try:
    from docx import Document
    import openpyxl
except ImportError:
    st.warning("Document libraries not available. Install: pip install python-docx openpyxl")

# Text Analysis
try:
    from wordcloud import WordCloud
    import seaborn as sns
except ImportError:
    st.warning("Text analysis libraries not available. Install: pip install wordcloud seaborn")

class FileProcessor:
    """Advanced file processing capabilities"""
    
    def __init__(self):
        self.supported_formats = {
            'images': ['.png', '.jpg', '.jpeg', '.gif', '.bmp', '.tiff'],
            'documents': ['.pdf', '.docx', '.txt', '.md'],
            'data': ['.csv', '.xlsx', '.json']
        }
    
    def process_uploaded_file(self, uploaded_file) -> Dict[str, Any]:
        """Process any uploaded file and extract information"""
        if uploaded_file is None:
            return {"error": "No file uploaded"}
        
        file_extension = os.path.splitext(uploaded_file.name)[1].lower()
        file_type = self._determine_file_type(file_extension)
        
        result = {
            "filename": uploaded_file.name,
            "file_type": file_type,
            "file_extension": file_extension,
            "file_size": uploaded_file.size,
            "processed_at": datetime.now().isoformat()
        }
        
        try:
            if file_type == "image":
                result.update(self._process_image(uploaded_file))
            elif file_type == "pdf":
                result.update(self._process_pdf(uploaded_file))
            elif file_type == "document":
                result.update(self._process_document(uploaded_file))
            elif file_type == "data":
                result.update(self._process_data_file(uploaded_file))
            else:
                result.update(self._process_text_file(uploaded_file))
                
        except Exception as e:
            result["error"] = f"Error processing file: {str(e)}"
        
        return result
    
    def _determine_file_type(self, extension: str) -> str:
        """Determine file type from extension"""
        for file_type, extensions in self.supported_formats.items():
            if extension in extensions:
                return file_type
        return "text"
    
    def _process_image(self, uploaded_file) -> Dict[str, Any]:
        """Process image files with OCR and analysis"""
        try:
            image = Image.open(uploaded_file)
            
            result = {
                "image_mode": image.mode,
                "image_size": image.size,
                "image_format": image.format
            }
            
            # Convert image to RGB if needed for OCR
            if image.mode != 'RGB':
                image = image.convert('RGB')
            
            # OCR text extraction
            try:
                extracted_text = pytesseract.image_to_string(image)
                result["extracted_text"] = extracted_text
                result["text_length"] = len(extracted_text)
                
                # Analyze extracted text
                if extracted_text.strip():
                    result["text_analysis"] = self._analyze_text(extracted_text)
                
            except Exception as e:
                result["ocr_error"] = f"OCR failed: {str(e)}"
            
            # Image analysis
            img_array = np.array(image)
            result["color_analysis"] = {
                "mean_rgb": img_array.mean(axis=(0,1)).tolist(),
                "brightness": img_array.mean(),
                "dominant_colors": self._get_dominant_colors(img_array)
            }
            
            return result
            
        except Exception as e:
            return {"error": f"Image processing failed: {str(e)}"}
    
    def _process_pdf(self, uploaded_file) -> Dict[str, Any]:
        """Process PDF files and extract content"""
        try:
            # Read PDF with pdfplumber for better text extraction
            with pdfplumber.open(uploaded_file) as pdf:
                text_content = ""
                page_count = len(pdf.pages)
                
                for page in pdf.pages:
                    text_content += page.extract_text() or ""
                
                result = {
                    "page_count": page_count,
                    "text_content": text_content,
                    "text_length": len(text_content),
                    "word_count": len(text_content.split()),
                }
                
                if text_content.strip():
                    result["text_analysis"] = self._analyze_text(text_content)
                    result["learning_topics"] = self._extract_learning_topics(text_content)
                
                return result
                
        except Exception as e:
            return {"error": f"PDF processing failed: {str(e)}"}
    
    def _process_document(self, uploaded_file) -> Dict[str, Any]:
        """Process Word documents and other text files"""
        try:
            if uploaded_file.name.endswith('.docx'):
                doc = Document(uploaded_file)
                text_content = "\n".join([paragraph.text for paragraph in doc.paragraphs])
            else:
                # Handle as text file
                text_content = str(uploaded_file.read(), "utf-8")
            
            result = {
                "text_content": text_content,
                "text_length": len(text_content),
                "word_count": len(text_content.split()),
                "paragraph_count": text_content.count('\n\n') + 1
            }
            
            if text_content.strip():
                result["text_analysis"] = self._analyze_text(text_content)
                result["learning_topics"] = self._extract_learning_topics(text_content)
            
            return result
            
        except Exception as e:
            return {"error": f"Document processing failed: {str(e)}"}
    
    def _process_data_file(self, uploaded_file) -> Dict[str, Any]:
        """Process data files (CSV, Excel, JSON)"""
        try:
            file_extension = os.path.splitext(uploaded_file.name)[1].lower()
            
            if file_extension == '.csv':
                df = pd.read_csv(uploaded_file)
            elif file_extension == '.xlsx':
                df = pd.read_excel(uploaded_file)
            elif file_extension == '.json':
                data = json.load(uploaded_file)
                df = pd.DataFrame(data) if isinstance(data, list) else pd.json_normalize(data)
            else:
                return {"error": "Unsupported data format"}
            
            result = {
                "data_shape": df.shape,
                "columns": df.columns.tolist(),
                "data_types": df.dtypes.to_dict(),
                "missing_values": df.isnull().sum().to_dict(),
                "summary_statistics": df.describe().to_dict() if df.select_dtypes(include=[np.number]).shape[1] > 0 else {},
                "sample_data": df.head().to_dict()
            }
            
            return result
            
        except Exception as e:
            return {"error": f"Data processing failed: {str(e)}"}
    
    def _process_text_file(self, uploaded_file) -> Dict[str, Any]:
        """Process plain text files"""
        try:
            text_content = str(uploaded_file.read(), "utf-8")
            
            result = {
                "text_content": text_content,
                "text_length": len(text_content),
                "word_count": len(text_content.split()),
                "line_count": text_content.count('\n') + 1
            }
            
            if text_content.strip():
                result["text_analysis"] = self._analyze_text(text_content)
                result["learning_topics"] = self._extract_learning_topics(text_content)
            
            return result
            
        except Exception as e:
            return {"error": f"Text processing failed: {str(e)}"}
    
    def _analyze_text(self, text: str) -> Dict[str, Any]:
        """Comprehensive text analysis"""
        words = text.split()
        sentences = text.split('.')
        
        # Basic statistics
        analysis = {
            "word_count": len(words),
            "sentence_count": len(sentences),
            "avg_words_per_sentence": len(words) / max(len(sentences), 1),
            "unique_words": len(set(words)),
            "lexical_diversity": len(set(words)) / max(len(words), 1)
        }
        
        # Find most common words
        word_freq = {}
        for word in words:
            word = word.lower().strip('.,!?";:')
            if len(word) > 3:
                word_freq[word] = word_freq.get(word, 0) + 1
        
        analysis["most_common_words"] = sorted(word_freq.items(), key=lambda x: x[1], reverse=True)[:10]
        
        # Educational content detection
        educational_keywords = [
            'definition', 'concept', 'theory', 'principle', 'method', 'process',
            'explain', 'understand', 'learn', 'study', 'analyze', 'calculate',
            'example', 'formula', 'equation', 'problem', 'solution', 'research'
        ]
        
        educational_score = sum(1 for word in words if word.lower() in educational_keywords)
        analysis["educational_content_score"] = educational_score / max(len(words), 1)
        
        return analysis
    
    def _extract_learning_topics(self, text: str) -> List[str]:
        """Extract potential learning topics from text"""
        # Simple topic extraction based on patterns
        topics = []
        
        # Look for definition patterns
        definition_patterns = [
            r'(\w+) is defined as',
            r'(\w+) refers to',
            r'(\w+) means',
            r'the concept of (\w+)',
            r'(\w+) theory',
            r'(\w+) principle'
        ]
        
        for pattern in definition_patterns:
            matches = re.findall(pattern, text, re.IGNORECASE)
            topics.extend(matches)
        
        # Remove duplicates and filter
        topics = list(set([topic.lower() for topic in topics if len(topic) > 3]))
        return topics[:20]  # Return top 20 topics
    
    def _get_dominant_colors(self, img_array: np.ndarray) -> List[List[int]]:
        """Extract dominant colors from image"""
        try:
            # Reshape image to list of pixels
            pixels = img_array.reshape(-1, 3)
            
            # Simple clustering to find dominant colors
            from sklearn.cluster import KMeans
            kmeans = KMeans(n_clusters=5, random_state=42)
            kmeans.fit(pixels)
            
            colors = kmeans.cluster_centers_.astype(int).tolist()
            return colors
        except:
            return []

class PDFGenerator:
    """Advanced PDF generation system"""
    
    def __init__(self):
        self.styles = getSampleStyleSheet() if 'getSampleStyleSheet' in globals() else None
    
    def generate_study_guide(self, topic: str, content: Dict[str, Any], user_profile: Dict) -> bytes:
        """Generate a comprehensive study guide PDF"""
        buffer = io.BytesIO()
        
        try:
            doc = SimpleDocTemplate(buffer, pagesize=letter)
            story = []
            
            # Title
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=self.styles['Heading1'],
                fontSize=24,
                spaceAfter=30,
                alignment=1  # Center alignment
            )
            story.append(Paragraph(f"Study Guide: {topic}", title_style))
            story.append(Spacer(1, 20))
            
            # User info
            story.append(Paragraph(f"Generated for: {user_profile.get('name', 'Student')}", self.styles['Normal']))
            story.append(Paragraph(f"Level: {user_profile.get('level', 'Intermediate')}", self.styles['Normal']))
            story.append(Paragraph(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", self.styles['Normal']))
            story.append(Spacer(1, 20))
            
            # Table of contents
            story.append(Paragraph("Table of Contents", self.styles['Heading2']))
            toc_data = [
                ["Section", "Page"],
                ["1. Introduction", "2"],
                ["2. Key Concepts", "3"],
                ["3. Examples", "4"],
                ["4. Practice Problems", "5"],
                ["5. Summary", "6"]
            ]
            
            toc_table = Table(toc_data)
            toc_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 14),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(toc_table)
            story.append(PageBreak())
            
            # Content sections
            sections = [
                ("Introduction", content.get('introduction', f"This study guide covers {topic} comprehensively.")),
                ("Key Concepts", content.get('key_concepts', "Important concepts will be listed here.")),
                ("Examples", content.get('examples', "Practical examples and applications.")),
                ("Practice Problems", content.get('practice_problems', "Problems to test understanding.")),
                ("Summary", content.get('summary', "Key takeaways and review points."))
            ]
            
            for section_title, section_content in sections:
                story.append(Paragraph(f"{section_title}", self.styles['Heading2']))
                story.append(Spacer(1, 12))
                story.append(Paragraph(section_content, self.styles['Normal']))
                story.append(Spacer(1, 20))
            
            doc.build(story)
            buffer.seek(0)
            return buffer.read()
            
        except Exception as e:
            # Fallback to simple PDF generation
            return self._generate_simple_pdf(topic, content, user_profile)
    
    def _generate_simple_pdf(self, topic: str, content: Dict[str, Any], user_profile: Dict) -> bytes:
        """Fallback simple PDF generation"""
        try:
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font("Arial", size=12)
            
            # Title
            pdf.set_font("Arial", 'B', 16)
            pdf.cell(200, 10, f"Study Guide: {topic}", ln=True, align='C')
            pdf.ln(10)
            
            # Content
            pdf.set_font("Arial", size=12)
            pdf.cell(200, 10, f"Generated for: {user_profile.get('name', 'Student')}", ln=True)
            pdf.cell(200, 10, f"Level: {user_profile.get('level', 'Intermediate')}", ln=True)
            pdf.cell(200, 10, f"Date: {datetime.now().strftime('%Y-%m-%d')}", ln=True)
            pdf.ln(10)
            
            # Add content
            for key, value in content.items():
                pdf.cell(200, 10, f"{key.title()}: {str(value)[:100]}...", ln=True)
            
            return pdf.output(dest='S').encode('latin1')
            
        except Exception as e:
            return f"Error generating PDF: {str(e)}".encode()
    
    def generate_quiz_pdf(self, quiz_data: Dict[str, Any], user_profile: Dict) -> bytes:
        """Generate a quiz as PDF"""
        buffer = io.BytesIO()
        
        try:
            doc = SimpleDocTemplate(buffer, pagesize=letter)
            story = []
            
            # Title
            title = f"Quiz: {quiz_data.get('subject', 'General Knowledge')}"
            story.append(Paragraph(title, self.styles['Heading1']))
            story.append(Spacer(1, 20))
            
            # Instructions
            instructions = """
            Instructions:
            1. Read each question carefully
            2. Choose the best answer for multiple choice questions
            3. Write True or False for true/false questions
            4. Show your work for calculation problems
            """
            story.append(Paragraph(instructions, self.styles['Normal']))
            story.append(Spacer(1, 20))
            
            # Questions
            for i, question in enumerate(quiz_data.get('questions', []), 1):
                # Question text
                q_text = f"Question {i}: {question.get('question', '')}"
                story.append(Paragraph(q_text, self.styles['Heading3']))
                
                # Options for multiple choice
                if question.get('type') == 'multiple_choice':
                    for j, option in enumerate(question.get('options', []), 1):
                        option_text = f"   {chr(96+j)}) {option}"
                        story.append(Paragraph(option_text, self.styles['Normal']))
                
                story.append(Spacer(1, 15))
            
            doc.build(story)
            buffer.seek(0)
            return buffer.read()
            
        except Exception as e:
            return self._generate_simple_quiz_pdf(quiz_data, user_profile)
    
    def _generate_simple_quiz_pdf(self, quiz_data: Dict[str, Any], user_profile: Dict) -> bytes:
        """Simple quiz PDF fallback"""
        try:
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font("Arial", size=12)
            
            # Title
            pdf.set_font("Arial", 'B', 16)
            title = f"Quiz: {quiz_data.get('subject', 'General Knowledge')}"
            pdf.cell(200, 10, title, ln=True, align='C')
            pdf.ln(10)
            
            # Questions
            pdf.set_font("Arial", size=12)
            for i, question in enumerate(quiz_data.get('questions', []), 1):
                pdf.cell(200, 10, f"Q{i}: {question.get('question', '')}", ln=True)
                
                if question.get('type') == 'multiple_choice':
                    for j, option in enumerate(question.get('options', []), 1):
                        pdf.cell(200, 10, f"   {chr(96+j)}) {option}", ln=True)
                
                pdf.ln(5)
            
            return pdf.output(dest='S').encode('latin1')
            
        except Exception as e:
            return f"Error generating quiz PDF: {str(e)}".encode()

class RoadmapGenerator:
    """Generate learning roadmaps and study plans"""
    
    def __init__(self):
        self.roadmap_templates = {
            "Mathematics": {
                "Beginner": ["Basic Arithmetic", "Algebra Basics", "Geometry Fundamentals", "Basic Statistics"],
                "Intermediate": ["Advanced Algebra", "Trigonometry", "Calculus I", "Probability"],
                "Advanced": ["Calculus II & III", "Linear Algebra", "Differential Equations", "Real Analysis"]
            },
            "Physics": {
                "Beginner": ["Classical Mechanics", "Basic Thermodynamics", "Waves and Sound", "Basic Electricity"],
                "Intermediate": ["Electromagnetism", "Modern Physics", "Optics", "Statistical Mechanics"],
                "Advanced": ["Quantum Mechanics", "Relativity", "Quantum Field Theory", "Condensed Matter"]
            },
            "Computer Science": {
                "Beginner": ["Programming Basics", "Data Structures", "Basic Algorithms", "Computer Systems"],
                "Intermediate": ["Algorithm Analysis", "Database Systems", "Software Engineering", "Computer Networks"],
                "Advanced": ["Machine Learning", "Distributed Systems", "Computer Graphics", "AI Research"]
            }
        }
    
    def generate_roadmap(self, subject: str, level: str, duration_weeks: int = 12) -> Dict[str, Any]:
        """Generate a personalized learning roadmap"""
        template = self.roadmap_templates.get(subject, {}).get(level, [])
        
        if not template:
            template = ["Foundation Concepts", "Core Principles", "Advanced Topics", "Applications"]
        
        weeks_per_topic = max(1, duration_weeks // len(template))
        
        roadmap = {
            "subject": subject,
            "level": level,
            "duration_weeks": duration_weeks,
            "topics": [],
            "generated_at": datetime.now().isoformat()
        }
        
        current_week = 1
        for i, topic in enumerate(template):
            topic_data = {
                "topic": topic,
                "week_start": current_week,
                "week_end": min(current_week + weeks_per_topic - 1, duration_weeks),
                "description": f"Learn and master {topic}",
                "goals": [
                    f"Understand core concepts of {topic}",
                    f"Complete practice problems",
                    f"Apply {topic} to real-world scenarios"
                ],
                "resources": [
                    "Textbook chapters",
                    "Online tutorials",
                    "Practice exercises",
                    "Video lectures"
                ]
            }
            roadmap["topics"].append(topic_data)
            current_week += weeks_per_topic
        
        return roadmap
    
    def create_roadmap_visualization(self, roadmap: Dict[str, Any]) -> go.Figure:
        """Create a visual roadmap using Plotly"""
        topics = roadmap["topics"]
        
        fig = go.Figure()
        
        # Create timeline
        for i, topic in enumerate(topics):
            fig.add_trace(go.Scatter(
                x=[topic["week_start"], topic["week_end"]],
                y=[i, i],
                mode='lines+markers',
                name=topic["topic"],
                line=dict(width=8),
                marker=dict(size=12)
            ))
            
            # Add topic labels
            fig.add_annotation(
                x=(topic["week_start"] + topic["week_end"]) / 2,
                y=i,
                text=topic["topic"],
                showarrow=False,
                font=dict(size=10, color="white"),
                bgcolor="rgba(0,0,0,0.5)",
                bordercolor="white",
                borderwidth=1
            )
        
        fig.update_layout(
            title=f"Learning Roadmap: {roadmap['subject']} ({roadmap['level']} Level)",
            xaxis_title="Week",
            yaxis_title="Topics",
            yaxis=dict(
                tickmode='array',
                tickvals=list(range(len(topics))),
                ticktext=[topic["topic"] for topic in topics]
            ),
            showlegend=False,
            height=400 + len(topics) * 50
        )
        
        return fig

class QuizFromFileGenerator:
    """Generate quizzes from uploaded files"""
    
    def __init__(self):
        self.question_types = ["multiple_choice", "true_false", "short_answer"]
    
    def generate_quiz_from_content(self, content: str, num_questions: int = 5, 
                                  subject: str = "General", difficulty: str = "Intermediate") -> Dict[str, Any]:
        """Generate quiz questions from text content"""
        if not content or len(content.strip()) < 100:
            return {"error": "Content too short to generate meaningful questions"}
        
        # Extract key information
        sentences = [s.strip() for s in content.split('.') if len(s.strip()) > 20]
        key_facts = self._extract_key_facts(content)
        
        questions = []
        
        for i in range(min(num_questions, len(key_facts))):
            if i < len(key_facts):
                question = self._create_question_from_fact(key_facts[i], difficulty)
                if question:
                    questions.append(question)
        
        # Fill remaining questions with general questions
        while len(questions) < num_questions and len(sentences) > 0:
            sentence = sentences.pop(0)
            question = self._create_question_from_sentence(sentence, difficulty)
            if question:
                questions.append(question)
        
        quiz = {
            "subject": subject,
            "difficulty": difficulty,
            "total_questions": len(questions),
            "questions": questions,
            "generated_from": "uploaded_content",
            "created_at": datetime.now().isoformat()
        }
        
        return quiz
    
    def _extract_key_facts(self, content: str) -> List[str]:
        """Extract key facts from content"""
        # Look for definition patterns
        patterns = [
            r'([A-Z][^.]*?) is ([^.]*\.)',
            r'([A-Z][^.]*?) means ([^.]*\.)',
            r'([A-Z][^.]*?) refers to ([^.]*\.)',
            r'The ([^.]*?) is ([^.]*\.)',
        ]
        
        facts = []
        for pattern in patterns:
            matches = re.findall(pattern, content)
            for match in matches:
                if len(match) == 2:
                    facts.append(f"{match[0]} is {match[1]}")
        
        return facts[:10]  # Return top 10 facts
    
    def _create_question_from_fact(self, fact: str, difficulty: str) -> Optional[Dict[str, Any]]:
        """Create a question from a key fact"""
        try:
            # Simple fact-based question
            if " is " in fact:
                parts = fact.split(" is ", 1)
                if len(parts) == 2:
                    subject_part = parts[0].strip()
                    definition_part = parts[1].strip()
                    
                    # Create multiple choice question
                    question = {
                        "question": f"What is {subject_part}?",
                        "type": "multiple_choice",
                        "options": [
                            definition_part,
                            "A different concept entirely",
                            "Not clearly defined",
                            "Under investigation"
                        ],
                        "correct_answer": definition_part,
                        "explanation": f"According to the content, {fact}"
                    }
                    return question
        except:
            pass
        
        return None
    
    def _create_question_from_sentence(self, sentence: str, difficulty: str) -> Optional[Dict[str, Any]]:
        """Create a question from a sentence"""
        if len(sentence.split()) < 5:
            return None
        
        # Create true/false question
        question = {
            "question": f"True or False: {sentence}",
            "type": "true_false",
            "options": ["True", "False"],
            "correct_answer": "True",
            "explanation": "This statement is directly from the provided content."
        }
        
        return question
